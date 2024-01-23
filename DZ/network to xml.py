import glob
import re
from ipaddress import IPv4Interface

from openpyxl import Workbook
# import os.path

wb = Workbook()
ws = wb.active
FILES_FOLDER = r"..\..\..\config_files\*.log"
FILE_XLSX = r"../../../config_files/network.xlsx"


def find_ip(string):
    if "ip address" in string:
        ip_plus_netmask = re.findall(
            r"((?:(?:25[0-5]|2[0-4][0-9]|1[0-9]{2}|[0-9]{2}|[0-9])(?:\.?|$)){4})",
            string
        )
        if len(ip_plus_netmask) == 2:
            ip_addr = ip_plus_netmask[0]+'/'+ip_plus_netmask[1]
            res = IPv4Interface(ip_addr)
            return res
    return None


result = []

for filename in glob.glob(FILES_FOLDER):
    with (open(filename) as fil):
        for string in fil:
            tango = find_ip(string.strip())
            if tango:
                result.append(tango.network)

ws['A1'] = "Network"
ws['B1'] = "Netmask"
i = 1

for net in set(result):
    i += 1
    field_A = 'A' + str(i)
    field_B = 'B' + str(i)
    ws[field_A] = str(net.network_address)
    ws[field_B] = str(net.netmask)

# if os.path.isfile(FILE_XLSX):
#    os.remove(FILE_XLSX)

wb.save(FILE_XLSX)
